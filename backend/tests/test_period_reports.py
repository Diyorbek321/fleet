"""Calendar-period reports — the monthly/weekly document.

The distinction from the rolling-window endpoints is the whole point: these
must return the same numbers to whoever opens them, whenever they open them.
Most of what is tested here is therefore about *which period a row falls into*,
because that is the only way a report can quietly disagree with itself.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta, timezone

import pytest
from httpx import AsyncClient

from app.services.period_reports import resolve_period


class TestPeriodBoundaries:
    def test_month_is_the_whole_calendar_month(self):
        p = resolve_period("month", 0, date(2026, 8, 15))
        assert (p.start, p.end) == (date(2026, 8, 1), date(2026, 8, 31))
        assert p.label == "2026 avgust"

    def test_previous_month_crosses_the_year(self):
        p = resolve_period("month", 1, date(2026, 1, 10))
        assert (p.start, p.end) == (date(2025, 12, 1), date(2025, 12, 31))

    def test_february_in_a_leap_year(self):
        p = resolve_period("month", 0, date(2028, 2, 5))
        assert p.end == date(2028, 2, 29)

    def test_week_runs_monday_to_sunday(self):
        # 2026-08-29 is a Saturday.
        p = resolve_period("week", 0, date(2026, 8, 29))
        assert p.start.weekday() == 0
        assert (p.start, p.end) == (date(2026, 8, 24), date(2026, 8, 30))

    def test_a_period_asked_for_twice_is_the_same_period(self):
        """A report re-run tomorrow must cover the same days as today.

        This is the property a rolling window cannot offer, and the reason
        these endpoints exist alongside it.
        """
        first = resolve_period("month", 1, date(2026, 8, 3))
        later = resolve_period("month", 1, date(2026, 8, 27))
        assert (first.start, first.end) == (later.start, later.end)

    def test_bounds_include_the_last_moment_of_the_last_day(self):
        """Built from midnight *after* the end, so 23:59 on the 31st is inside.

        Taking the end as 23:59:59 instead would silently drop the final minute
        of the month — and month-end is exactly when deliveries cluster.
        """
        p = resolve_period("month", 0, date(2026, 8, 15))
        start, end = p.bounds_utc()
        last_moment = datetime(2026, 8, 31, 23, 59, 59, tzinfo=timezone.utc)
        assert start <= last_moment < end + timedelta(hours=5)


async def _truck(client: AsyncClient, admin_headers, name="T1", plate="01 A 111 AA") -> str:
    res = await client.post(
        "/api/trucks", headers=admin_headers, json={"name": name, "plate_number": plate}
    )
    assert res.status_code == 200, res.text
    return res.json()["id"]


class TestPeriodReportEndpoint:
    async def test_an_empty_period_reports_zeroes_not_an_error(
        self, client: AsyncClient, admin_headers
    ):
        """A quiet month is a legitimate answer, and the document still exists."""
        res = await client.get("/api/reports/period?kind=month&offset=2", headers=admin_headers)
        assert res.status_code == 200, res.text
        body = res.json()
        assert body["trips_delivered"] == 0
        assert body["revenue"] == 0
        assert body["profit"] == 0
        assert body["margin_pct"] == 0

    async def test_revenue_is_recognised_when_the_load_is_delivered(
        self, client: AsyncClient, admin_headers
    ):
        truck_id = await _truck(client, admin_headers)
        trip = await client.post(
            "/api/trips", headers=admin_headers, json={"truck_id": truck_id, "rate": 5_000_000}
        )
        assert trip.status_code == 200, trip.text
        tid = trip.json()["id"]
        for to_status in ("en_route", "delivered"):
            await client.post(
                f"/api/trips/{tid}/advance", headers=admin_headers, json={"to_status": to_status}
            )

        body = (
            await client.get("/api/reports/period?kind=month&offset=0", headers=admin_headers)
        ).json()
        assert body["trips_delivered"] == 1
        assert body["revenue"] == 5_000_000

    async def test_a_running_trip_is_counted_apart_not_dropped(
        self, client: AsyncClient, admin_headers
    ):
        """Revenue in flight has to be visible, or the report looks like a loss."""
        truck_id = await _truck(client, admin_headers, "T2", "01 A 222 BB")
        await client.post(
            "/api/trips", headers=admin_headers, json={"truck_id": truck_id, "rate": 9_000_000}
        )

        body = (
            await client.get("/api/reports/period?kind=month&offset=0", headers=admin_headers)
        ).json()
        assert body["trips_in_progress"] >= 1
        assert body["revenue"] == 0  # not earned yet

    async def test_fuel_and_maintenance_land_in_the_period_they_were_paid(
        self, client: AsyncClient, admin_headers
    ):
        truck_id = await _truck(client, admin_headers, "T3", "01 A 333 CC")
        await client.post(
            f"/api/trucks/{truck_id}/fuel-logs",
            headers=admin_headers,
            json={"liters": 100, "cost_per_liter": 13_000},
        )
        await client.post(
            f"/api/trucks/{truck_id}/maintenance",
            headers=admin_headers,
            json={
                "service_type": "oil_change",
                "cost": 500_000,
                "performed_at": date.today().isoformat(),
            },
        )

        body = (
            await client.get("/api/reports/period?kind=month&offset=0", headers=admin_headers)
        ).json()
        assert body["fuel_cost"] == 1_300_000
        assert body["maintenance_cost"] == 500_000
        assert body["total_cost"] == 1_800_000
        assert body["profit"] == -1_800_000

    async def test_consumption_is_withheld_for_a_week(self, client: AsyncClient, admin_headers):
        """Litres bought is not litres burned over a span this short.

        The same tank fills read 36.5 L/100km across a month and 13.6 and 110.6
        across two weeks inside it. Printing any of the short-period figures
        invites a decision about a driver who did nothing different.
        """
        res = await client.get("/api/reports/period?kind=week&offset=0", headers=admin_headers)
        assert res.status_code == 200
        assert res.json()["consumption_reliable"] is False

    async def test_the_spreadsheet_is_a_real_workbook(self, client: AsyncClient, admin_headers):
        res = await client.get(
            "/api/reports/period.xlsx?kind=month&offset=0", headers=admin_headers
        )
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert "hisobot-oylik-" in res.headers["content-disposition"]
        # PK zip magic: proves a workbook was written, not an error page.
        assert res.content[:2] == b"PK"

        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(res.content))
        assert wb.sheetnames == ["Xulosa", "Mashinalar", "Haydovchilar"]

    async def test_money_cells_are_numbers_not_text(self, client: AsyncClient, admin_headers):
        """A column of strings that look like money cannot be summed.

        Which is the first thing the accountant this file exists for will do.
        """
        truck_id = await _truck(client, admin_headers, "T4", "01 A 444 DD")
        await client.post(
            f"/api/trucks/{truck_id}/fuel-logs",
            headers=admin_headers,
            json={"liters": 10, "cost_per_liter": 13_000},
        )
        res = await client.get(
            "/api/reports/period.xlsx?kind=month&offset=0", headers=admin_headers
        )

        from io import BytesIO

        from openpyxl import load_workbook

        ws = load_workbook(BytesIO(res.content))["Mashinalar"]
        fuel_cell = ws.cell(row=2, column=5)
        assert isinstance(fuel_cell.value, (int, float))
        assert fuel_cell.number_format == "#,##0"

    async def test_a_bad_period_kind_is_refused(self, client: AsyncClient, admin_headers):
        res = await client.get("/api/reports/period?kind=quarter", headers=admin_headers)
        assert res.status_code == 422

    async def test_requires_authentication(self, client: AsyncClient):
        assert (await client.get("/api/reports/period?kind=month")).status_code == 401


class TestPeriodReportIsTenantScoped:
    async def test_another_orgs_money_is_not_in_your_report(self, client: AsyncClient):
        async def signup(email: str, org: str) -> dict[str, str]:
            await client.post(
                "/api/auth/register",
                json={"email": email, "password": "password123", "org_name": org},
            )
            login = await client.post(
                "/api/auth/login", json={"email": email, "password": "password123"}
            )
            return {"Authorization": f"Bearer {login.json()['access_token']}"}

        a = await signup("pa@org.com", "Period Org A")
        b = await signup("pb@org.com", "Period Org B")

        truck_id = await _truck(client, a, "A-Truck", "01 A 999 ZZ")
        await client.post(
            f"/api/trucks/{truck_id}/fuel-logs",
            headers=a,
            json={"liters": 1000, "cost_per_liter": 13_000},
        )

        body = (await client.get("/api/reports/period?kind=month&offset=0", headers=b)).json()
        assert body["fuel_cost"] == 0
        assert body["trucks"] == []
