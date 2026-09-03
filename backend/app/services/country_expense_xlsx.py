"""Render the country-expense breakdown as a spreadsheet.

Three sheets, because there are three questions people bring to this report and
one grid cannot answer all of them:

* **Davlatlar** — the whole range rolled up: what each country cost, itemised.
  The screen answers this too; the sheet exists so it can be pasted into a
  workbook next to last quarter's.
* **Reyslar** — one row per round trip, three countries across the columns.
  This is the sheet that gets sorted: which run was expensive, and where.
* **Tafsilot** — every line of every trip, one per row. Nothing is summarised,
  so anything the first two sheets hint at can be traced to its cell.

Native amounts and USD sit in separate columns and are never mixed: a column of
tenge added to a column of roubles is a number with no meaning. Where no
exchange rate was available the USD cell is left genuinely empty rather than
zeroed, so a blank reads as "not converted" and never as "cost nothing".
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.country_expenses import CountryExpenseReport

_HEAD_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=15)
_MUTED_FONT = Font(color="6B7280", size=9)
_TOTAL_FONT = Font(bold=True)

MONEY = "#,##0.00"
# So'm and tenge are never quoted with a fractional part in practice.
WHOLE_MONEY = "#,##0"
LITRES = "#,##0.0"
KM = "#,##0.0"
RATE = "#,##0.00"

COUNTRY_LABELS = {"uz": "O'zbekiston", "kz": "Qozog'iston", "ru": "Rossiya"}

# The driver's form is in Russian, so the categories keep the words the person
# who filled it in actually used — a translated label makes a dispatcher check
# whether it means the same line.
CATEGORY_LABELS = {
    "fuel": "Yoqilg'i (GSM)",
    "insurance": "Sug'urta",
    "platon": "Платон",
    "food": "Ovqat",
    "traffic_police": "ГАИ",
    "adblue": "AdBlue",
    "fine": "Jarima",
    "spare_parts": "Zapchast",
    "repair": "Ta'mir",
    "refund": "Qaytim",
    "parking": "Stoyanka",
    "phone": "Telefon",
    "transport": "Transport",
    "shower": "Dush",
    "groceries": "Oziq-ovqat",
    "parking_paperwork": "Stoyanka rasmiylashtirish",
    "taxi": "Taksi",
    "carwash": "Moyka",
}

RATE_SOURCE_LABELS = {
    "trip": "reys kursi",
    "org": "tashkilot kursi",
    "mixed": "aralash",
}

CARD_LABELS = {"doha": "DOHA karta", "e1card": "E1 karta"}


def _country_label(code: str) -> str:
    return COUNTRY_LABELS.get(code, code.upper())


def _category_label(code: str) -> str:
    return CATEGORY_LABELS.get(code, code)


def _headers(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _money(ws, row: int, col: int, value, fmt: str = MONEY):
    """Write a number, or leave the cell empty when there is none.

    ``None`` means "no rate, nothing converted". Writing 0 there would put a
    figure in a total column that someone will sum.
    """
    if value is None:
        return None
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    return cell


def _title(ws, report: CountryExpenseReport, subtitle: str) -> int:
    scope = report.organization or "Fleet"
    ws["A1"] = f"{scope} — davlatlar bo'yicha reys xarajatlari"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Davr: {report.start:%d.%m.%Y} – {report.end:%d.%m.%Y} · {subtitle}"
    ws["A3"] = f"Tayyorlandi: {report.generated_at:%d.%m.%Y %H:%M} UTC"
    ws["A3"].font = _MUTED_FONT
    return 5


def _sheet_countries(ws, report: CountryExpenseReport) -> None:
    row = _title(ws, report, f"{len(report.trips)} ta reys")
    _widths(ws, [30, 16, 16, 14, 16])

    for block in report.countries:
        if block.total == 0 and block.fuel_liters == 0:
            continue
        ws.cell(row=row, column=1, value=_country_label(block.country)).font = _TOTAL_FONT
        note = RATE_SOURCE_LABELS.get(block.rate_source or "", "kurs yo'q")
        if block.rate:
            ws.cell(row=row, column=2, value=f"1 USD = {block.rate:,.2f} {block.currency} ({note})")
        else:
            ws.cell(row=row, column=2, value="Kurs kiritilmagan — USD hisoblanmadi")
        ws.cell(row=row, column=2).font = _MUTED_FONT
        row += 1

        _headers(ws, row, ["Nima uchun", block.currency, "USD", "Litr", "Ulushi"])
        row += 1

        for line in block.lines:
            ws.cell(row=row, column=1, value=_category_label(line.category))
            _money(ws, row, 2, line.amount, WHOLE_MONEY)
            _money(ws, row, 3, line.amount_usd, MONEY)
            _money(ws, row, 4, line.liters, LITRES)
            if block.total:
                share = ws.cell(row=row, column=5, value=line.amount / block.total)
                share.number_format = "0.0%"
            row += 1

        ws.cell(row=row, column=1, value="Jami").font = _TOTAL_FONT
        cell = _money(ws, row, 2, block.total, WHOLE_MONEY)
        if cell:
            cell.font = _TOTAL_FONT
        cell = _money(ws, row, 3, block.total_usd, MONEY)
        if cell:
            cell.font = _TOTAL_FONT
        cell = _money(ws, row, 4, block.fuel_liters or None, LITRES)
        if cell:
            cell.font = _TOTAL_FONT
        row += 2

    if report.cards:
        ws.cell(row=row, column=1, value="Kartadagi yoqilg'i (davlatga taqsimlanmaydi)").font = _TOTAL_FONT
        row += 1
        _headers(ws, row, ["Karta", "Summa", "", "Litr", ""])
        row += 1
        for card in report.cards:
            ws.cell(row=row, column=1, value=CARD_LABELS.get(card.column, card.column))
            _money(ws, row, 2, card.amount, MONEY)
            _money(ws, row, 4, card.liters, LITRES)
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Jami (USD)").font = _TOTAL_FONT
    cell = _money(ws, row, 3, report.total_usd, MONEY)
    if cell:
        cell.font = _TOTAL_FONT
    if report.usd_partial:
        missing = ", ".join(_country_label(c) for c in report.countries_missing_rate)
        ws.cell(
            row=row + 1,
            column=1,
            value=f"Diqqat: {missing} uchun kurs yo'q — USD jami to'liq emas.",
        ).font = _MUTED_FONT


def _sheet_trips(ws, report: CountryExpenseReport) -> None:
    row = _title(ws, report, "har bir reys bir qator")
    header_row = row

    labels = ["Reys", "Mashina", "Haydovchi", "Sana", "Yo'nalish", "km"]
    for block_country in ("uz", "kz", "ru"):
        currency = {"uz": "UZS", "kz": "KZT", "ru": "RUB"}[block_country]
        labels += [f"{_country_label(block_country)} ({currency})", f"{_country_label(block_country)} (USD)"]
    labels.append("Jami (USD)")
    _headers(ws, header_row, labels)
    _widths(ws, [16, 18, 18, 12, 30, 10, 16, 14, 16, 14, 16, 14, 14])
    row += 1

    for trip in report.trips:
        ws.cell(row=row, column=1, value=trip.reference)
        ws.cell(row=row, column=2, value=trip.plate_number or trip.truck_name or "—")
        ws.cell(row=row, column=3, value=trip.driver_name or "—")
        if trip.trip_date:
            cell = ws.cell(row=row, column=4, value=trip.trip_date)
            cell.number_format = "DD.MM.YYYY"
        ws.cell(row=row, column=5, value=trip.route or "—")
        _money(ws, row, 6, trip.distance_km, KM)

        by_country = {b.country: b for b in trip.countries}
        for i, code in enumerate(("uz", "kz", "ru")):
            block = by_country.get(code)
            if block is None:
                continue
            _money(ws, row, 7 + i * 2, block.total or None, WHOLE_MONEY)
            _money(ws, row, 8 + i * 2, block.total_usd, MONEY)
        _money(ws, row, 13, trip.total_usd, MONEY)
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _sheet_lines(ws, report: CountryExpenseReport) -> None:
    row = _title(ws, report, "har bir xarajat qatori")
    header_row = row
    _headers(
        ws,
        header_row,
        ["Reys", "Mashina", "Sana", "Davlat", "Nima uchun", "Summa", "Valyuta", "USD", "Litr", "Kurs manbai"],
    )
    _widths(ws, [16, 18, 12, 16, 26, 16, 10, 14, 10, 16])
    row += 1

    for trip in report.trips:
        for block in trip.countries:
            for line in block.lines:
                ws.cell(row=row, column=1, value=trip.reference)
                ws.cell(row=row, column=2, value=trip.plate_number or trip.truck_name or "—")
                if trip.trip_date:
                    cell = ws.cell(row=row, column=3, value=trip.trip_date)
                    cell.number_format = "DD.MM.YYYY"
                ws.cell(row=row, column=4, value=_country_label(block.country))
                ws.cell(row=row, column=5, value=_category_label(line.category))
                _money(ws, row, 6, line.amount, WHOLE_MONEY)
                ws.cell(row=row, column=7, value=block.currency)
                _money(ws, row, 8, line.amount_usd, MONEY)
                _money(ws, row, 9, line.liters, LITRES)
                ws.cell(
                    row=row,
                    column=10,
                    value=RATE_SOURCE_LABELS.get(block.rate_source or "", "kurs yo'q"),
                )
                row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_workbook(report: CountryExpenseReport) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Davlatlar"
    _sheet_countries(ws, report)

    _sheet_trips(wb.create_sheet("Reyslar"), report)
    _sheet_lines(wb.create_sheet("Tafsilot"), report)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename_for(report: CountryExpenseReport) -> str:
    """A name that says which range — and which truck — the file holds."""
    scope = "reys-xarajatlari"
    if report.truck_id:
        plate = next((t.plate_number for t in report.trips if t.plate_number), None)
        if plate:
            safe = "".join(ch if ch.isalnum() else "-" for ch in plate).strip("-")
            scope = f"{scope}-{safe}"
    return f"{scope}-{report.start:%Y%m%d}-{report.end:%Y%m%d}.xlsx"
