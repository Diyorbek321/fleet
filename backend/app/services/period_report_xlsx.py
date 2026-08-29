"""Render a period report as a spreadsheet.

Excel rather than PDF because of who opens it. The owner reads the summary on
the screen the report came from; the person who needs a *file* is the
accountant, and they need to sort it, filter it and paste it into their own
workbook. A PDF of the same numbers is a picture of data they would retype.

Values are written as numbers with a display format, never as pre-formatted
strings — a column of text that looks like money cannot be summed, which is the
first thing anyone does with it.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.period_reports import PeriodReport

_HEAD_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=15)
_TOTAL_FONT = Font(bold=True)
_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))

# So'm has no minor unit in practice — nobody invoices in tiyin.
MONEY = "#,##0"
LITRES = "#,##0.0"
KM = "#,##0.0"
COUNT = "#,##0"


def _headers(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook(report: PeriodReport) -> bytes:
    wb = Workbook()

    # ── Xulosa ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Xulosa"

    ws["A1"] = f"{report.organization} — {report.period.label}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Davr: {report.period.start:%d.%m.%Y} – {report.period.end:%d.%m.%Y}"
    ws["A3"] = f"Tayyorlandi: {report.generated_at:%d.%m.%Y %H:%M} UTC"
    ws["A3"].font = Font(color="6B7280", size=9)

    # Costs are written negative so the column adds up to the profit line on
    # its own. An accountant checking the arithmetic should not have to know
    # which rows to subtract.
    rows: list[tuple[str, object, str]] = [
        ("Yetkazilgan reyslar", report.trips_delivered, COUNT),
        ("Jarayondagi reyslar", report.trips_in_progress, COUNT),
        ("", "", ""),
        ("Daromad", report.revenue, MONEY),
        ("Yoqilg'i xarajati", -report.fuel_cost, MONEY),
        ("Haydovchi xarajatlari", -report.expense_cost, MONEY),
        ("Texnik xizmat", -report.maintenance_cost, MONEY),
        ("Jami xarajat", -report.total_cost, MONEY),
        ("Sof foyda", report.profit, MONEY),
        ("Rentabellik, %", report.margin_pct, "#,##0.0"),
        ("", "", ""),
        ("Bosib o'tilgan yo'l, km", report.distance_km, KM),
        ("Yoqilg'i, litr", report.fuel_liters, LITRES),
        # Suppressed rather than shown with a footnote: a number in a cell
        # gets copied into someone else's workbook, and the caveat does not
        # travel with it.
        (
            "Sarfiyot, L/100km",
            report.l_per_100km if report.consumption_reliable else "ma'lumot yetarli emas",
            "#,##0.0" if report.consumption_reliable else "General",
        ),
    ]

    row = 5
    for label, value, fmt in rows:
        if not label:
            row += 1
            continue
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        if label in ("Sof foyda", "Jami xarajat"):
            ws.cell(row=row, column=1).font = _TOTAL_FONT
            cell.font = _TOTAL_FONT
        row += 1

    if not report.consumption_reliable:
        ws.cell(
            row=row + 1,
            column=1,
            value=(
                "Sarfiyot ko'rsatilmadi: quyilgan yoqilg'i yoqilgan yoqilg'i emas — "
                "bakda qolgani keyingi davrga o'tadi. Qisqa davrda bu farq natijani "
                "buzadi. Oylik hisobotga yoki Leakage sahifasiga qarang: u har bir "
                "mashinani flot medianasiga solishtiradi."
            ),
        ).font = Font(color="B45309", italic=True)
        row += 1

    if report.distance_partial:
        # Said on the sheet, not only in the API response: whoever opens the
        # file is the one who would otherwise divide fuel by a distance that
        # covers part of the period and read the answer as consumption.
        ws.cell(
            row=row + 1,
            column=1,
            value=(
                "Diqqat: davrning bir qismi GPS saqlash muddatidan eski. "
                "Masofa va L/100km to'liq emas."
            ),
        ).font = Font(color="B45309", italic=True)

    _widths(ws, [30, 20])

    # ── Mashinalar ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Mashinalar")
    _headers(ws, 1, [
        "Mashina", "Davlat raqami", "Reys", "Daromad", "Yoqilg'i",
        "Haydovchi xarajati", "Texnik xizmat", "Jami xarajat", "Foyda",
        "Yo'l, km", "Litr", "L/100km",
    ])
    for i, line in enumerate(report.trucks, start=2):
        values = [
            line.name, line.plate_number, line.trips, line.revenue, line.fuel_cost,
            line.expense_cost, line.maintenance_cost, line.total_cost, line.profit,
            line.distance_km, line.fuel_liters, line.l_per_100km,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = _BORDER
            if col == 3:
                cell.number_format = COUNT
            elif col in (4, 5, 6, 7, 8, 9):
                cell.number_format = MONEY
            elif col == 10:
                cell.number_format = KM
            elif col in (11, 12):
                cell.number_format = LITRES
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(2, len(report.trucks) + 1)}"
    _widths(ws, [22, 16, 8, 16, 16, 18, 16, 16, 16, 12, 12, 11])

    # ── Haydovchilar ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Haydovchilar")
    _headers(ws, 1, ["Haydovchi", "Reys", "Daromad", "Xarajat"])
    for i, driver in enumerate(report.drivers, start=2):
        values = [driver.name, driver.trips, driver.revenue, driver.expense_cost]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = _BORDER
            if col == 2:
                cell.number_format = COUNT
            elif col in (3, 4):
                cell.number_format = MONEY
    ws.freeze_panes = "A2"
    _widths(ws, [26, 8, 18, 18])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def filename_for(report: PeriodReport) -> str:
    """A filename that sorts chronologically in a folder full of them."""
    kind = "oylik" if report.period.kind == "month" else "haftalik"
    return f"hisobot-{kind}-{report.period.start:%Y-%m-%d}.xlsx"
